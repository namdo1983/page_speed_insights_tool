import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment
import speedtest


class MyUtils:
    def __init__(self, driver, chrome_options, excel_path, file_name):
        self.driver = driver
        self.chrome_options = chrome_options
        self.excel_path = excel_path


    def my_open_chrome_browser(self):
        # declare option chrome
        self.chrome_options.add_argument("--disable-popup-blocking")
        self.chrome_options.add_argument("--disable-notifications")
        self.chrome_options.add_argument("--incognito")
        self.chrome_options.add_argument("log-level=3")
        self.driver.set_window_position(0, 0)
        self.driver.set_window_size(1360, 768)
        self.driver.implicitly_wait(30)
        self.driver.set_page_load_timeout(60)
        self.driver.delete_all_cookies()


    def write_excel_result_performance(self, result_row, num_col, result):
        book = openpyxl.load_workbook(self.excel_path)
        sheet = book.active
        sheet.cell(row=result_row, column=num_col, value=result)
        book.save(self.excel_path)


    def fill_color(self):
        wb = openpyxl.load_workbook(self.excel_path)
        ws = wb.active
        mr = ws.max_row
        mc = ws.max_column
        red_fill = PatternFill(start_color="ff0000",
                               end_color="ff0000", fill_type="solid")
        green_fill = PatternFill(
            start_color="00cc00", end_color="00cc00", fill_type="solid"
        )
        yellow_fill = PatternFill(
            start_color="00FFFF00", end_color="00FFFF00", fill_type="solid"
        )
        gray_fill = PatternFill(
            start_color="00808080", end_color="00808080", fill_type="solid"
        )
        lightblue_fill = PatternFill(
            start_color="00CCFFCC", end_color="00CCFFCC", fill_type="solid"
        )
        orange_fill = PatternFill(
            start_color="00FF9900", end_color="00FF9900", fill_type="solid"
        )
        # fill color for result response
        for col in range(4, mc + 1):
            for row in range(2, mr + 1):
                if ws.cell(row, col).value is not None:
                    if "OK" in ws.cell(row, col).value:
                        ws.cell(row, col).fill = green_fill
                    elif ws.cell(row, col).value.strip().startswith('This site'):
                        ws.cell(row, col).fill = red_fill
                    elif ws.cell(row, col).value.strip().startswith('Error Code'):
                        ws.cell(row, col).fill = orange_fill

        # color score red(0-49), orange(50-89), green(90-100)
        for col in range(2, mc):
            for row in range(2, mr + 1):
                # print('Max col: ', mc)
                if ws.cell(row, col).value is not None:
                    if 0 < ws.cell(row, col).value <= 49:
                        ws.cell(row, col).fill = red_fill
                    if 50 <= ws.cell(row, col).value <= 89:
                        ws.cell(row, col).fill = orange_fill
                    if ws.cell(row, col).value >= 90:
                        ws.cell(row, col).fill = green_fill

        # center all text sheet["C1"].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        for col in range(1, mc + 1):
            for row in range(1, mr + 1):
                ws.cell(row, col).alignment = Alignment(
                    horizontal='center', vertical='center', wrap_text=True)
                ws.cell(row, 1).fill = lightblue_fill

        # define border All
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))
        for col in range(1, mc + 1):
            for row in range(1, mr + 1):
                ws.cell(row, col).border = thin_border
                # fill header value
                if ws.cell(row, col).value is None:
                    continue
                else:
                    ws.cell(1, col).fill = yellow_fill
        wb.save(self.excel_path)


    def check_network(self):
        try:
            s = speedtest.Speedtest()
            res = s.get_config()
            # ip_nm = res["client"]["ip"]
            ten_nm = res["client"]["isp"]
            return ten_nm
        except:
            print('Can not get ISP.')


    def create_excel(self):
        book = openpyxl.Workbook()
        sheet = book.active
        sheet["A1"] = "Domain"
        sheet.column_dimensions["A"].width = 40
        sheet["B1"] = "Mobile Score"
        sheet.column_dimensions["B"].width = 10
        sheet["C1"] = "Desktop Score"
        sheet.column_dimensions["C"].width = 10
        book.save(self.excel_path)
        print(f'The "{self.excel_path}" created successful.')


    def write_excel(self, domain_row, domain):
        book = openpyxl.load_workbook(self.excel_path)
        sheet = book.active
        nha_mang = self.check_network()
        try:
            if "Viettel" in nha_mang:
                sheet["D1"] = "Viettel"
                sheet.column_dimensions["D"].width = 20
            elif "VNPT" in nha_mang:
                sheet["D1"] = 'VNPT'
                sheet.column_dimensions["D"].width = 20
            elif "FPT" in nha_mang:
                sheet["D1"] = 'FPT'
                sheet.column_dimensions["D"].width = 20

            # write domain at start row =2
            sheet[f"A{domain_row}"] = domain
            # print("[*] Domain =>", domain)
        except:
            print('Something wrong went write data to excel sheet.')

        book.save(self.excel_path)


    def write_excel_result(self, result_row, result):
        book = openpyxl.load_workbook(self.excel_path)
        sheet = book.active
        nha_mang = self.check_network()
        try:
            # write domain at start row =2
            if "Viettel" in nha_mang:
                sheet[f"D{result_row}"] = result
            elif "VNPT" in nha_mang:
                sheet[f"D{result_row}"] = result
            elif "FPT" in nha_mang:
                sheet[f"D{result_row}"] = result
        except:
            print('Something wrong went write data to excel sheet.')
        book.save(self.excel_path)


    def read_file_txt(self, file_name):
        urls = []
        if '.xlsx' in file_name:
            assert not isinstance(file_name,
                                  type(None)), f'Excel {file_name} not found.\nPlease check data.xlsx in current path.'

            wb = openpyxl.load_workbook(file_name)
            # ws = wb['Links']
            ws = wb.active
            data2 = []
            for row in ws.iter_rows(min_row=2, max_col=1, max_row=ws.max_row):
                for cell in row:
                    try:
                        if cell.value is not None:
                            val = cell.value
                            data2.append(val)
                    except TypeError as te:
                        print(str(te))

            for url in data2:
                urls.append(url.rstrip())
        else:
            with open(file_name, 'r', encoding='utf-8') as f:
                data = f.read().splitlines()
                for url in data:
                    if url.strip() is not '':  # remove white space in file txt
                        urls.append(url.rstrip())

        print(urls)
        return urls
